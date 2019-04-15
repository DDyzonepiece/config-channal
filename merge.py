
import openpyxl
from openpyxl.styles import *
from openpyxl.utils import *

#合并单元格函数file_path是文件路径，aa是需要合并的列索引列表,
#sheet_index是需要处理的列表索引不给的话默认第一个
def merge_cellr(file_path, aa, sheet_index=0):
	#加载已经存在的excel
	workbook = openpyxl.load_workbook(file_path) 
	# workbook = openpyxl.Workbook(path)
	name_list = workbook.sheetnames
	# worksheet = workbook.get_sheet_by_name(name_list[0])  #最新版本已经不能使用这种方法
	worksheet = workbook[name_list[sheet_index]]
	num_c=len(aa)
	print(num_c)
	num_r=worksheet.max_row
	for i in aa:
		unempty_index=[]
		
		for j in range(num_r):
			if worksheet.cell(row=j+1,column=i+1).value!=None:
				unempty_index.append(j+1)
				
		#print(empty_index)
		for k in range(len(unempty_index)-1):
			if unempty_index[k+1]-unempty_index[k]>1:
				worksheet.merge_cells(start_row=unempty_index[k], start_column=i+1, 
				end_row=unempty_index[k+1]-1, end_column=i+1)
		if num_r!=unempty_index[-1]:
			worksheet.merge_cells(start_row=unempty_index[-1], start_column=i+1, 
				end_row=num_r, end_column=i+1)
	
	#print(worksheet['D4'].value)
	#worksheet.merge_cells(start_row=2, start_column=1, end_row=13, end_column=1)
	#将需要合并单元格的列数转换成字符串
	bb=[]
	for i in aa:
		temp=get_column_letter(i+1)
		bb.append(temp)
	print(bb)
	
	
	
	
	#调整合并完单元格的位置
	for i in bb:
		#调整单元格列宽
		
		#worksheet.column_dimensions[i].auto_size = True
		
		#worksheet.column_dimensions[i].autofit= True
		
		
		a1=worksheet[i]
		#print(worksheet.columns)
		alignment1=Alignment(horizontal='center',
						vertical='center',
						text_rotation=0,
						wrap_text=False,
						shrink_to_fit=False,
						indent=0)
		for cell in a1:
			cell.alignment = alignment1
	#a1.c=alignment1
	#print(a1.value)
	workbook.save('merge.xlsx')

file_path=r'config.xlsx'
aa=[0,1,2,3,4,5,6]

merge_cellr(file_path, aa)
print(111)


